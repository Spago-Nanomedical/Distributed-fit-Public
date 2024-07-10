# -*- coding: utf-8 -*-
"""
Created on Thu Mar  9 10:26:13 2023
Program to do a modelbased fit to plasma concentrations from
clinical study SPAGOPIX-01.
A numerical simulation based on the known distribution of renal
filtration pore sizes and the known distribution of particle sizes.
The renal pore sizes are modeled as following the logistic equation centered
at 5.5 nm and with a slope based on known filtration rates of various
proteins. Each size is filtered by the kidneys according to their GFR which
corresponds to kMax. THe blood volume is calculated from the weight of the
patient and the GFR is estimated from age.The particles are assumed to
have a Gaussian size distribution when injected:
    c(r,t)=a*exp(-(x-b)^2/(2c^2)
where a = doseConcentration/(sigmaSize*sqrt(2*pi)) based on the integral
of a Gaussian distribution. The integral should be
equal to the doseConcentration.
,b is the average size
and c is the standard deviation (square root of variance).
The injected dose is then subjected to a size dependent clearance according
to known and estimated properties of the renal system. The clearance is
scaled according to the sigmoid curve, the "logistic equation":
kMax/(1+exp(filterSlope(x-filterCutoff))) where
kMax is the excretion rate for a small polar compound with 100% renal clearance
filterSlope is how sharp the filter cuts
filterCutoff is where the inflection point is, should correspond to 50%
clearance.
The particles have an adjustable number of pools of various sizes and
correspondingly for the filter.
The simulation takes infusion time into account and for each size pool it
is excreted by a simple first order mechanism. Integration is done with a
constant timestep
@author: oskar.axelsson

Version 2.0 saves all figures as .png files and writes all patient data in
an excel file. It runs the program 14 times (once for each patient).

Version 2.2 Calculates the GFR as well as possible based on age, height, 
and weight. Due to GDPR concerns all body measures of the patients have been set to default
"""


import numpy as np
from matplotlib import pyplot
from openpyxl import load_workbook
from openpyxl import Workbook
import uuid

# import matplotlib Activate if Font size in plots needs to be changed
from scipy.optimize import curve_fit
import time


class Patient:
    """Describes a patient and the data"""

    def __init__(self):
        self.ID = ''
        self.bodyWeight = 70.0
        self.height = 165.0
        self.age = 60
        self.nominalDose = 0.0  # umol/kg
        self.tPts = np.array([])
        self.conc = np.array([])  # mM!!!
        self.GFR = 92.0 #ml/min
        self.cMax = 0.0 #uM
        self.AUC = 0.0 #uM x min
        self.actualDose = 0.0 #umol
        self.R2 = 0.0
        
    def calcGFR(self):
    # Relative average GFR for women aged 50 to 90: GFR = 148.3 -1.018 * age
    # Adapted from: https://doi.org/10.1681/asn.2020020151
        relGFR = 148.3 - 1.018 * self.age
    # Body area accoring to the DuBois formula, 
    # DuBois D, DuBois EF. A formula to estimate the approximate surface
    # area if height and weight be known. 
    # Arch Intern Medicine. 1916; 17:863-871
        self.GFR = relGFR * 0.007184 * self.bodyWeight**0.425 \
        * self.height**0.725 / 1.73
    
    def calcCMax(self, cData):
        self.cMax = max(cData)
        
    def calcAUC(self, cData):
        self.AUC = np.trapz(cData)
        
    def calcR2(self):
        avgC = np.mean(self.conc)
        cSpread = self.conc-avgC
        cSq = [c**2 for c in cSpread]
        cVar = sum(cSq)
        residualsSq = []
        fitPoints = distributedFiltration(self.tPts, self.actualDose)
        for i in range(len(self.tPts)):
            residualsSq.append((self.conc[i]-fitPoints[i])**2)
        resSqSum = sum(residualsSq)
        self.R2 = 1 - resSqSum/cVar

def initPatients():
    myPatients = []

    patient1 = Patient()
    patient1.ID = 'Patient 1'
    patient1.nominalDose = 10 * 13 * patient1.bodyWeight  # Si/Mn is 13, umol
    patient1.tPts = np.array([0, 10, 20, 61, 65, 75, 90, 105, 161, 540, 1500, 14400])
    patient1.conc = np.array([0.023, 0.35, 0.69, 2.03, 1.858, 1.793, 1.577, 1.403, 1.117, 0.515, 0.305, 0.0854])
    myPatients.append(patient1)

    patient2 = Patient()
    patient2.ID = 'Patient 2'
    patient2.nominalDose = 10 * 13 * patient2.bodyWeight  # Si/Mn is 13
    patient2.tPts = np.array([0, 10, 20, 60, 65, 75, 90, 105, 179, 540, 1500, 23040])
    patient2.conc = np.array([0.0266, 0.352, 0.788, 1.982, 2.025, 1.851, 1.746, 1.488, 1.098, 0.624, 0.355, 0.0389])
    myPatients.append(patient2)

    patient3 = Patient()
    patient3.ID = 'Patient 3'
    patient3.nominalDose = 10 * 13 * patient3.bodyWeight  # Si/Mn is 13
    patient3.tPts = np.array([0, 10, 20 , 60, 65, 75, 90, 105, 170, 540, 1500, 17280])
    patient3.conc = np.array([0.034, 0.293, 0.642, 1.516, 1.484, 1.360, 1.183, 1.033, 0.808, 0.448, 0.279, 0.0543])
    myPatients.append(patient3)

    patient4 = Patient()
    patient4.ID = 'Patient 4'
    patient4.nominalDose = 10 * 13 * patient4.bodyWeight  # Si/Mn is 13
    patient4.tPts = np.array([0, 10, 20, 62, 65, 75, 90, 105, 170, 555, 1450, 5760])
    patient4.conc = np.array([0.018, 0.395, 0.728, 1.992, 1.921, 1.787, 1.515, 1.413, 1.169, 0.555, 0.379, 0.157])
    myPatients.append(patient4)

    patient5 = Patient()
    patient5.ID = 'Patient 5'
    patient5.nominalDose = 10 * 13 * patient5.bodyWeight  # Si/Mn is 13
    patient5.tPts = np.array([0, 10, 20, 63, 65, 75, 90, 105, 170, 542, 1455, 17280])
    patient5.conc = np.array([0.0189, 0.450, 0.873, 1.920, 1.983, 1.734, 1.531, 1.407, 1.080, 0.536, 0.347, 0.150])
    myPatients.append(patient5)

    patient6 = Patient()
    patient6.ID = 'Patient 6'
    patient6.nominalDose = 10 * 13 * patient6.bodyWeight  # Si/Mn is 13
    patient6.tPts = np.array([0, 10, 20, 60, 66, 75, 90, 105, 173, 540, 1455, 25920])
    patient6.conc = np.array([0.023, 0.366, 0.587, 1.483, 1.680, 1.498, 1.294, 1.183, 0.9414, 0.5113, 0.336, 0.0436])
    myPatients.append(patient6)

    patient7 = Patient()
    patient7.ID = 'Patient 7'
    patient7.nominalDose = 20 * 13 * patient7.bodyWeight  # Si/Mn is 13
    patient7.tPts = np.array([0, 10, 20, 60, 67, 75, 90, 105, 175, 540, 1501, 25930])
    patient7.conc = np.array([0.0156, 0.634, 1.430, 3.139, 2.960, 2.801, 2.515, 2.029, 1.726, 0.899, 0.808, 0.0654])
    myPatients.append(patient7)

    patient8 = Patient()
    patient8.age = 71
    patient8.nominalDose = 20 * 13 * patient8.bodyWeight  # Si/Mn is 13
    patient8.tPts = np.array([0, 10, 20, 60, 65, 75, 90, 105, 160, 545, 1500, 5700])
    patient8.conc = np.array([0, 0.387, 1.104, 2.866, 2.926, 2.612, 2.290, 2.135, 1.711, 0.976, 0.582, 0.257])
    myPatients.append(patient8)

    patient9 = Patient()
    patient9.ID = 'Patient 9'
    patient9.nominalDose = 20 * 13 * patient9.bodyWeight  # Si/Mn is 13
    patient9.tPts = np.array([0, 10, 19, 61, 65, 77, 90, 105, 180, 540, 1495, 7125])
    patient9.conc = np.array([0.0379, 0.797, 1.523, 4.446, 4.499, 3.837, 3.512, 3.114, 2.656, 1.253, 0.703, 0.255])
    myPatients.append(patient9)

    patient10 = Patient()
    patient10.ID = 'Patient 10'
    patient10.nominalDose = 20 * 13 * patient10.bodyWeight  # Si/Mn is 13
    patient10.tPts = np.array([0, 10, 20, 60, 65, 75, 90, 105, 161, 537, 1490, 5588])
    patient10.conc = np.array([0.0,  0.679, 1.207, 3.077, 3.009, 2.774, 2.428, 2.311, 1.914, 0.992, 0.562, 0.282])
    myPatients.append(patient10)

    patient11 = Patient()
    patient11.ID = 'Patient 11'
    patient11.nominalDose = 20 * 13 * patient11.bodyWeight  # Si/Mn is 13
    patient11.tPts = np.array([0, 10, 20, 60, 66, 90, 105, 158, 543, 8667])
    patient11.conc = np.array([0.0, 0.881, 1.513, 3.756, 3.454, 2.923, 2.622, 2.236, 1.039, 0.194])
    myPatients.append(patient11)

    patient12 = Patient()
    patient12.ID = 'Patient 12'
    patient12.nominalDose = 20 * 13 * patient12.bodyWeight  # Si/Mn is 13
    patient12.tPts = np.array([0, 11, 20, 61, 65, 75, 90, 105, 176, 540, 9964])
    patient12.conc = np.array([0.049, 0.751, 1.431, 3.568, 3.554, 3.106, 2.895, 2.572, 2.416, 1.139, 0.170])
    myPatients.append(patient12)    

    patient13 = Patient()
    patient13.ID = 'Patient 13'
    patient13.nominalDose = 20 * 13 * patient13.bodyWeight  # Si/Mn is 13
    patient13.tPts = np.array([0, 10, 20, 60, 65, 75, 90, 103, 180, 541, 5590])
    patient13.conc = np.array([0.0, 0.627, 1.341, 3.537, 3.311, 2.936, 2.631, 2.431, 1.900, 1.018, 0.279])
    myPatients.append(patient13)

    patient14 = Patient()
    patient14.ID = 'Patient 14'
    patient14.nominalDose = 20 * 13 * patient14.bodyWeight  # Si/Mn is 13
    patient14.tPts = np.array([0, 10, 20, 60, 65, 75, 90, 106, 180, 545, 2700])
    patient14.conc = np.array([0.0, 0.291, 0.970, 2.397, 2.435, 2.200, 1.879, 1.727, 1.398, 0.775, 0.334]) #mM
    myPatients.append(patient14)
    

    for i in range(len(myPatients)):
        #Adjust from mM to uM
        myPatients[i].conc *= 1000
        
        myPatients[i].calcGFR() 

    return myPatients


def doDistributed(myPatient):
    startTime = time.time()
    # curve fit function to optimize the actual dose to fit experimental data
    # (function to call, x-values, y-values, initial guess of parameter(s))
    optVar, pCovariance = curve_fit(distributedFiltration, myPatient.tPts,
                                    myPatient.conc,  p0=(myPatient.nominalDose))
    # Optimized parameter:
    doseOpt = optVar[0]  # umol/kg
    # Create a smooth fit curve
    # For comparable AUC it is better to use the same large max t
    maxT = 30000
    t = myPatient.tPts
    t_line = np.arange(t[1], maxT, 1)
    y_line = distributedFiltration(t_line, doseOpt)
    plotCurveFit(t_line, y_line)
    myPatient.calcCMax(y_line)
    myPatient.calcAUC(y_line)
    myPatient.actualDose = doseOpt
    myPatient.calcR2()
    printResults(myPatient)
    writeResultsToExcel(myPatient)
    exeTime = time.time() - startTime
    print('ExecutionTime = ', exeTime)
    

def plotCurveFit(t_line, y_line):
    pyplot.figure(1)
    pyplot.clf()
    pyplot.plot(t_line, y_line, '-g')
    pyplot.plot(currentPatient.tPts, currentPatient.conc, linestyle='', marker='o')
    pyplot.xscale('log')
    pyplot.xlabel('min')
    pyplot.ylabel('uM Si')
    pyplot.title(currentPatient.ID)
    pyplot.savefig(str(currentPatient.ID)+"-datapoints.png")
    pyplot.show()

def printResults(myPatient):
    print(' ')
    print(myPatient.ID)
    print('cMax =', myPatient.cMax, 'uM')
    print('AUC=', myPatient.AUC, 'uM x min')
    print('Nominal dose:', myPatient.nominalDose, 'umol,', myPatient.nominalDose / myPatient.bodyWeight, 'umol/kg')
    print('Actual dose', myPatient.actualDose, 'umol', myPatient.actualDose / myPatient.bodyWeight, 'umol/kg')
    print('Body mass:', myPatient.bodyWeight)
    print('GFR:', myPatient.GFR)
    print('Goodness of fit, R2 =', myPatient.R2)
    
def writeResultsToExcel(myPatient):  
    rowNr = i + 5 #The i from main is used
    sheet.cell(row=rowNr, column=2).value = myPatient.ID
    sheet.cell(row=rowNr, column=3).value = myPatient.bodyWeight
    sheet.cell(row=rowNr, column=4).value = myPatient.GFR
    sheet.cell(row=rowNr, column=5).value = myPatient.cMax
    sheet.cell(row=rowNr, column=6).value = myPatient.AUC
    sheet.cell(row=rowNr, column=7).value = myPatient.nominalDose
    sheet.cell(row=rowNr, column=8).value = myPatient.actualDose
    sheet.cell(row=rowNr, column=9).value = myPatient.R2
    
def distributedFiltration(t, dose):
    """A numerical simulation based on the known distribution of renal
    filtration pore sizes and the known distribution of particle sizes.
    THe renal pore sizes are modeled as following the logistic equation centered
    at 5.5 nm and with a slope based on known filtration rates of various
    proteins. Each size is filtered by the kidneys according to their GFR which
    corresponds to kMax. THe blood volume is calculated from the weight of the
    patient and the GFR is estimated from age.The particles are assumed to
    have a Gaussian size distribution when injected:
        c(r,t)=a*exp(-(x-b)^2/(2c^2)
    where a = doseConcentration/(sigmaSize*sqrt(2*pi)) based on the integral
    of a Gaussian distribution. The integral should be
    equal to the doseConcentration.
    ,b is the average size
    and c is the standard deviation (square root of variance).
    The injected dose is then subjected to a size dependent clearance according
    to known and estimated properties of the renal system. The clearance is
    scaled according to the sigmoid curve, the "logistic equation":
    kMax/(1+exp(filterSlope(x-filterCutoff))) where
    kMax is the excretion rate for a small polar compound with 100% renal clearance
    filterSlope is how sharp the filter cuts
    filterCutoff is where the inflection point is, should correspond to 50%
    clearance.
    The particles have an adjustable number of pools of various sizes and
    correspondingly for the filter.
    The simulation takes infusion time into account and for each size pool it
    is excreted by a simple first order mechanism. Integration is done with a
    constant timestep"""

    global NoFunctionCalls
    # Simulation parameters
    deltaT = 1.0  # time resolution min
    # print(t)
    time =  np.arange(0, max(t)+1, deltaT)  # one extra to catch last experimental timePt
    noTimePts = len(time)
    sizeResolution = 50  # How fine grained the size profile is. 30 seems enough

    # Physiology parameters

    filtrationCutOff = 5.5  # nm
    filtrationSlope = 3.8  # /nm
    bodyMass = currentPatient.bodyWeight  # kg Taken from main because the opt function call is so messy
    hematocrit = 0.42
    infusionTime = 60  # min
    plasmaVolume = bodyMass * 0.0613 * (1 - hematocrit) * 1000  # ml, checked

    """Blood fraction from
    https://www.omnicalculator.com/health/blood-volume 
    Average hematokrit from https://www.medicinenet.com/hematocrit/article.htm"""
    # GFR for different ages from https://www.disabled-world.com/health/cancer/kidney/gfr-ckd.php
    # GFR = 90 ml/min #Estimate for age around 60
    kMax = currentPatient.GFR / plasmaVolume  # (ml/min) / ml = /min

    # Drug parameters
    print('Dose', dose, 'uM')
    print('Nominal plasma concentration', currentPatient.nominalDose
          / plasmaVolume * 1000, 'uM')  # uM
    print('Max possible conc, no clearance', dose / plasmaVolume * 1000, 'uM')
    avgSize = 5.8  # nm diameter 5.5-5.6 best estimate from DLS, 5.8 from GPC
    sigmaSize = 0.84  # nm
    minSize = avgSize - 3 * sigmaSize  # nm
    maxSize = avgSize + 3 * sigmaSize  # nm
    particleSize = np.linspace(minSize, maxSize, sizeResolution)  # nm, checked
    # particleSize = np.logspace(np.log10(minSize), np.log10(maxSize), sizeResolution)
    dX = (maxSize-minSize) / sizeResolution
    # print('Particle size)', particleSize)
    # create an array of amounts of each size
    a = dose / (sigmaSize * np.sqrt(2 * np.pi))  # umol/nm
    # print('a:', a)
    # Normal distributed sizes:
    expArr = np.exp(-0.5 * ((particleSize - avgSize) / sigmaSize)**2)
    # logNormal distributed sizes:
    # expArr = np.exp(-0.5 * ((np.log(particleSize) - np.log(avgSize)) / np.log(sigmaSize))**2)
    sizeDistrDose = np.array([a * e for e in expArr]) # umol/nm
    # print('SizeDistrDose', sizeDistrDose)
    fig = pyplot.figure(2)
    pyplot.plot(particleSize, sizeDistrDose)
    pyplot.xticks(range(int(np.floor(minSize)), int(np.ceil(maxSize))))
    pyplot.title('Particle size dose (umol)')
    pyplot.show


    #print('Sum dose', sum(sizeDistrDose) * dX, 'umol')
    #print('Dose', dose, 'umol')
    sizeDistrDoseRate = sizeDistrDose / infusionTime  # umol/(min nm)
    #print('Dose rate sum', sum(sizeDistrDoseRate) * dX, 'umol/min')
    #print('Dose rate:', dose/infusionTime, 'umol/min')

    # Create an array of rate constants, one for each size
    k = kMax / (1 + np.exp(filtrationSlope * (particleSize - filtrationCutOff)))  # min^-1

    # Initiation of arrays
    # This will be huge. Up to 26000 * 20 = 520 000 numbers, 4 bytes / float
    # = 2 GB. Should be fine.
    conc = np.zeros([noTimePts, sizeResolution])
    # Start integration, skip the last time point since we use i+1 index
    for i in range(noTimePts-1):
        # Are we still infusing?
        if time[i] <= infusionTime:
            k1 = 1
        else:
            k1 = 0
        # dC will be an array

        dCin = (k1 * sizeDistrDoseRate / (plasmaVolume / 1000)) * deltaT  # uM
        # print('dCin', dCin)
        dCout = np.array([-k[j] * conc[i][j] * deltaT for j in range(sizeResolution)])
        # print('dCout', dCout)
        dC = dCin + dCout
        # print('dC', dC)
        conc[i + 1] = conc[i] + dC

    concentrations = np.array([sum(c) * dX for c in conc])

    reportConc = []
    for i in range(noTimePts):
        if time[i] in t:
            reportConc.append(concentrations[i])
            # print('Timepts', time[i])
    NoFunctionCalls += 1
    print('iteration', NoFunctionCalls)

    return reportConc 


def biexp(t, conc0, weight1, rate1, rate2):
    
    """Biexponential function"""
    #The two exponentials have different weights and the weights sum to 1
    weight2 = 1 - weight1
    fValue = conc0 * (weight1 * np.exp(-rate1 * t) + weight2 * np.exp(-rate2 * t))
    return fValue


def doBiexp(myPatient):
    """Run a biexponential fit"""
    # curve fit function to optimize parameterss to fit experimental data
    t = myPatient.tPts[3:]
    expConc = myPatient.conc[3:] # uM
    popt, pCovariance = curve_fit(biexp, t, expConc,  p0=[3000.0, 0.5, 0.1, 0.01])
    #Optimized parameters:
    conc0, weight1, rate1, rate2 = popt
    #Create a smooth fit curve
    t_line = np.arange(min(t), max(t), 1)
    t_line[0] = 0.1  #For log plotting
    y_line = biexp(t_line, conc0, weight1, rate1, rate2)
    pyplot.plot(t_line, y_line, '-r')
    pyplot.show()
    print(patientId)
    print('Starting concentration =', conc0)
    print('Weight of rate constant 1 =', weight1)
    print('Rate constant 1 =', rate1)
    print('Rate constant 2 =', rate2)
    print('Half-life 1 =', np.log(2)/rate1)
    print('Half-life 2 =', np.log(2)/rate2)
    #Calculate R2
    avgC = np.mean(expConc)
    cSpread = expConc-avgC
    cSq = [c**2 for c in cSpread]
    cVar = sum(cSq)
    residualsSq = []
    for i in range(len(t)):
        residualsSq.append( (expConc[i]-biexp(t[i], conc0, weight1, rate1, rate2))**2)
    resSqSum = sum(residualsSq)
    R2 = 1 - resSqSum/cVar
    print('Goodness of fit, R2 =', R2)


def monoexp(t, conc0, rate):
    """Monoexponential function"""
    fValue = conc0 * np.exp(-rate * t)
    return fValue


def doMonoexp(myPatient):
    """Run a monoexponential fit"""
    # curve fit function to optimize parameterss to fit experimental data
    t = myPatient.tPts[3:]
    expConc = myPatient.conc[3:] # uM
    popt, pCovariance = curve_fit(monoexp, t, expConc,  p0=[3000.0, 0.1])
    #Optimized parameters:
    conc0, rate = popt
    #Create a smooth fit curve
    t_line = np.arange(min(t), max(t), 1)
    t_line[0] = 0.1  #For log plotting
    y_line = monoexp(t_line, conc0, rate)
    pyplot.plot(t_line, y_line, '-r')
    pyplot.show()
    print(patientId)
    print('Starting concentration =', conc0)
    print('Rate constant =', rate)
    print('Half-life =', np.log(2)/rate)
    #Calculate R2
    avgC = np.mean(expConc)
    cSpread = expConc-avgC
    cSq = [c**2 for c in cSpread]
    cVar = sum(cSq)
    residualsSq = []
    for i in range(len(t)):
        residualsSq.append( (expConc[i]-monoexp(t[i], conc0, rate))**2)
    resSqSum = sum(residualsSq)
    R2 = 1 - resSqSum/cVar
    print('Goodness of fit, R2 =', R2)

def setupExcelSheetForOutput():    
    wb = Workbook()
    sheet = wb.active
    
    # Write Excel headers
    sheet.cell(row=4, column=2).value = 'Patient Id'
    sheet.cell(row=4, column=3).value = 'Body mass (kg)'
    sheet.cell(row=4, column=4).value = 'Calc. GFR absolute \n (ml/min)'
    sheet.cell(row=4, column=5).value = 'Cmax Si (uM)'
    sheet.cell(row=4, column=6).value = 'AUC Si (uM min)'
    sheet.cell(row=4, column=7).value = 'Nominal dose Si (umol/kg)'
    sheet.cell(row=4, column=8).value = 'Actual dose Si (umol)'
    sheet.cell(row=4, column=9).value = 'R2'
    
    return wb

def saveExcelOutput(myWorkBook):
    urlID = str(uuid.uuid4())  #Generates a random ID number, giving each file a unique identifier
    excelURL ="C:\AddYourPathHere"+urlID+".xlsx"
    myWorkBook.save(filename=excelURL)
    
    
myWorkBook = setupExcelSheetForOutput()
sheet = myWorkBook.active
myPatients = initPatients()

for i in range(len(myPatients)):
    currentPatient = myPatients[i]
    NoFunctionCalls = 0 # Counter for the curve fitting optimization

    method = 'Distributed'
    if method == 'Distributed':
        doDistributed(currentPatient)
    if method == 'Monoexp':
        doMonoexp(currentPatient)
    if method == 'Biexp':
        doBiexp(currentPatient)

saveExcelOutput(myWorkBook)
 
